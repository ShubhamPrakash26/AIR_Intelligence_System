"""Unit tests for Week 11/12 output formatters.

Covers MarkdownFormatter, ExcelFormatter, and APSANewsletterFormatter against
synthetic EditorialReport and APSAArticle instances.
No LLM or network calls are made.
"""

from __future__ import annotations

import io

import pytest

from src.insights.editorial_models import APSAArticle, EditorialReport, EditorialSection
from src.insights.formatters import APSANewsletterFormatter, ExcelFormatter, MarkdownFormatter


# ---------------------------------------------------------------------------
# Sample data
# ---------------------------------------------------------------------------


def _make_section(
    theme: str = "root_cause",
    title: str = "Latent Conditions in Airway Preparation",
    narrative: str = (
        "Three incidents share a latent failure in pre-induction documentation. "
        "In each case no Plan B was recorded before anaesthesia commenced, reflecting "
        "an organisational norm of relying on individual experience."
    ),
    key_learning: str = "Pre-induction Plan A/B documentation must be a mandatory checklist item.",
    citations: list[str] | None = None,
) -> EditorialSection:
    return EditorialSection(
        section_id="sec-001",
        theme=theme,
        title=title,
        narrative=narrative,
        supporting_insights=[],
        evidence_citations=citations if citations is not None else [
            "Incident abc001 | severity=High | type=Airway Event | score=0.91",
            "Incident abc002 | severity=Moderate | type=Airway Event | score=0.78",
        ],
        key_learning=key_learning,
        tone_score=1.0,
        generated_at="2026-06-14T00:00:00+00:00",
        model_version="week10-editorial-engine-0.1",
    )


def _make_report(
    sections: list[EditorialSection] | None = None,
    evocative_title: str = "",
    clinical_references: list[str] | None = None,
) -> EditorialReport:
    secs = sections if sections is not None else [
        _make_section("root_cause", "Root Cause Findings"),
        _make_section(
            "safety_recommendations",
            title="System-Level Safety Actions",
            citations=["Incident abc003 | severity=High | score=0.85"],
        ),
    ]
    all_cits: list[str] = []
    seen: set[str] = set()
    for s in secs:
        for c in s.evidence_citations:
            if c not in seen:
                all_cits.append(c)
                seen.add(c)

    return EditorialReport(
        report_id="rep-001",
        query="recurring airway safety patterns",
        title="Clinical Safety Analysis: recurring airway safety patterns",
        executive_summary=(
            "Five incidents were reviewed. Two thematic areas emerge: latent documentation "
            "gaps in pre-induction planning and the absence of mandatory positional-change "
            "reassessment protocols during gynaecological laparoscopy."
        ),
        sections=secs,
        conclusion=(
            "The incidents reviewed reflect systemic conditions that a reporting culture "
            "is designed to surface. The actions identified are achievable at department level."
        ),
        total_incidents_referenced=5,
        all_citations=all_cits,
        tone_score=1.0,
        has_llm_narrative=True,
        generated_at="2026-06-14T00:00:00+00:00",
        model_version="week10-editorial-engine-0.1",
        evocative_title=evocative_title,
        clinical_references=clinical_references if clinical_references is not None else [],
    )


def _make_apsa_article(
    title: str = "AIRWAY EVENTS: THE LATENT GAPS IN PREPAREDNESS",
    vignette: str = "A high-severity airway event was reported during a laparoscopic procedure.",
    body: str = "Airway events remain one of the most consequential complications in anaesthesia.\n\nSystemic latent conditions allow these events to recur.",
    references: list[str] | None = None,
) -> APSAArticle:
    return APSAArticle(
        article_id="art-001",
        incident_id="inc-abc001",
        title=title,
        vignette=vignette,
        body=body,
        clinical_references=references if references is not None else [
            "1. Cook TM, et al. Major complications of airway management in the UK. Br J Anaesth. 2011;106(5):617-631.",
        ],
        incident_metadata={"severity": "High", "incident_type": ["Airway Event"]},
        generated_at="2026-06-14T00:00:00+00:00",
    )


# ---------------------------------------------------------------------------
# MarkdownFormatter — APSA format
# ---------------------------------------------------------------------------


class TestMarkdownFormatter:
    def test_format_returns_string(self):
        md = MarkdownFormatter().format(_make_report())
        assert isinstance(md, str)

    def test_title_in_output(self):
        report = _make_report()
        md = MarkdownFormatter().format(report)
        assert "Clinical Safety Analysis" in md

    def test_evocative_title_used_when_present(self):
        report = _make_report(evocative_title="SILENT GAPS: THE RECURRENCE OF AIRWAY EVENTS")
        md = MarkdownFormatter().format(report)
        assert "SILENT GAPS: THE RECURRENCE OF AIRWAY EVENTS" in md
        assert "# SILENT GAPS" in md

    def test_fallback_title_when_no_evocative_title(self):
        report = _make_report(evocative_title="")
        md = MarkdownFormatter().format(report)
        assert "Clinical Safety Analysis" in md

    def test_executive_summary_content_present(self):
        md = MarkdownFormatter().format(_make_report())
        assert "Five incidents were reviewed" in md

    def test_section_titles_as_bold_headers(self):
        md = MarkdownFormatter().format(_make_report())
        assert "**Root Cause Findings**" in md
        assert "**System-Level Safety Actions**" in md

    def test_narratives_present(self):
        md = MarkdownFormatter().format(_make_report())
        assert "latent failure in pre-induction documentation" in md

    def test_conclusion_present(self):
        md = MarkdownFormatter().format(_make_report())
        assert "reporting culture" in md

    def test_no_key_learning_blockquote(self):
        md = MarkdownFormatter().format(_make_report())
        assert "> **Key Learning:**" not in md

    def test_no_supporting_evidence_list(self):
        md = MarkdownFormatter().format(_make_report())
        assert "**Supporting Evidence:**" not in md

    def test_no_section_h2_headers(self):
        md = MarkdownFormatter().format(_make_report())
        assert "## Executive Summary" not in md
        assert "## Conclusion" not in md

    def test_clinical_references_section_when_present(self):
        refs = [
            "1. Cook TM, et al. Major complications of airway management. Br J Anaesth. 2011;106(5):617-631.",
            "2. Apfelbaum JL, et al. Practice guidelines for management of the difficult airway. Anesthesiology. 2013;118(2):251-270.",
        ]
        report = _make_report(clinical_references=refs)
        md = MarkdownFormatter().format(report)
        assert "**References**" in md
        assert "Cook TM" in md
        assert "Apfelbaum" in md

    def test_incident_evidence_shown_when_no_clinical_references(self):
        report = _make_report(clinical_references=[])
        md = MarkdownFormatter().format(report)
        assert "**Supporting Incident Evidence**" in md
        assert "abc001" in md

    def test_no_references_section_when_no_citations_at_all(self):
        report = EditorialReport(
            report_id="rep-empty",
            query="empty",
            title="T",
            executive_summary="No incidents for this query.",
            sections=[],
            conclusion="No patterns identified.",
            total_incidents_referenced=0,
            all_citations=[],
            tone_score=1.0,
            has_llm_narrative=False,
            generated_at="2026-06-14T00:00:00+00:00",
            model_version="v0",
        )
        md = MarkdownFormatter().format(report)
        assert "**References**" not in md
        assert "**Supporting Incident Evidence**" not in md

    def test_horizontal_rules_separate_major_sections(self):
        md = MarkdownFormatter().format(_make_report())
        assert md.count("---") >= 3

    def test_published_date_in_header(self):
        md = MarkdownFormatter().format(_make_report())
        assert "**Published:**" in md

    def test_footer_present(self):
        md = MarkdownFormatter().format(_make_report())
        assert "AIR Clinical Incident Intelligence Engine" in md

    def test_empty_sections_report_renders(self):
        report = EditorialReport(
            report_id="rep-empty",
            query="empty query",
            title="Clinical Safety Analysis: empty query",
            executive_summary="No incidents retrieved for this query.",
            sections=[],
            conclusion="No patterns identified.",
            total_incidents_referenced=0,
            all_citations=[],
            tone_score=1.0,
            has_llm_narrative=False,
            generated_at="2026-06-14T00:00:00+00:00",
            model_version="week10-editorial-engine-0.1",
        )
        md = MarkdownFormatter().format(report)
        assert isinstance(md, str)
        assert "No incidents retrieved" in md
        assert "No patterns identified" in md


# ---------------------------------------------------------------------------
# ExcelFormatter
# ---------------------------------------------------------------------------


class TestExcelFormatter:
    def test_format_returns_bytes(self):
        result = ExcelFormatter().format(_make_report())
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_is_valid_xlsx(self):
        import openpyxl
        data = ExcelFormatter().format(_make_report())
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_summary_sheet_present(self):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(_make_report())))
        assert "Summary" in wb.sheetnames

    def test_section_sheets_created(self):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(_make_report())))
        assert len(wb.sheetnames) >= 3  # Summary + 2 sections + Citations

    def test_citations_sheet_present_when_citations_exist(self):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(_make_report())))
        assert "Citations" in wb.sheetnames

    def test_no_citations_sheet_when_empty(self):
        import openpyxl
        report = EditorialReport(
            report_id="rep-nc",
            query="q",
            title="T",
            executive_summary="S" * 35,
            sections=[],
            conclusion="C" * 35,
            total_incidents_referenced=0,
            all_citations=[],
            tone_score=1.0,
            has_llm_narrative=False,
            generated_at="2026-06-14T00:00:00+00:00",
            model_version="v0",
        )
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        assert "Citations" not in wb.sheetnames

    def test_summary_sheet_contains_title(self):
        import openpyxl
        report = _make_report()
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        ws = wb["Summary"]
        cell_values = [str(ws.cell(row=r, column=2).value or "") for r in range(1, 20)]
        assert any("Clinical Safety Analysis" in v for v in cell_values)

    def test_summary_sheet_contains_executive_summary(self):
        import openpyxl
        report = _make_report()
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        ws = wb["Summary"]
        all_text = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, 30)
        )
        assert "Five incidents were reviewed" in all_text

    def test_section_sheet_contains_narrative(self):
        import openpyxl
        report = _make_report()
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        root_sheet = next(
            (name for name in wb.sheetnames if "Root" in name), None
        )
        assert root_sheet is not None
        ws = wb[root_sheet]
        all_text = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, 30)
        )
        assert "latent failure" in all_text

    def test_section_sheet_contains_key_learning(self):
        import openpyxl
        report = _make_report()
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        root_sheet = next(n for n in wb.sheetnames if "Root" in n)
        ws = wb[root_sheet]
        all_text = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, 30)
        )
        assert "Plan A/B" in all_text

    def test_section_sheet_name_max_31_chars(self):
        import openpyxl
        long_theme = "a_very_long_theme_name_that_exceeds_the_excel_limit"
        section = _make_section(theme=long_theme)
        report = _make_report(sections=[section])
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        for name in wb.sheetnames:
            assert len(name) <= 31

    def test_citations_sheet_content(self):
        import openpyxl
        report = _make_report()
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        ws = wb["Citations"]
        all_text = " ".join(
            str(ws.cell(row=r, column=2).value or "") for r in range(1, 20)
        )
        assert "abc001" in all_text

    def test_empty_report_produces_summary_only(self):
        import openpyxl
        report = EditorialReport(
            report_id="rep-e",
            query="q",
            title="T",
            executive_summary="S" * 35,
            sections=[],
            conclusion="C" * 35,
            total_incidents_referenced=0,
            all_citations=[],
            tone_score=1.0,
            has_llm_narrative=False,
            generated_at="2026-06-14T00:00:00+00:00",
            model_version="v0",
        )
        wb = openpyxl.load_workbook(io.BytesIO(ExcelFormatter().format(report)))
        assert "Summary" in wb.sheetnames
        assert len(wb.sheetnames) == 1

    def test_base64_round_trip(self):
        import base64
        import openpyxl
        report = _make_report()
        raw_bytes = ExcelFormatter().format(report)
        b64 = base64.b64encode(raw_bytes).decode()
        decoded = base64.b64decode(b64)
        wb = openpyxl.load_workbook(io.BytesIO(decoded))
        assert "Summary" in wb.sheetnames


# ---------------------------------------------------------------------------
# APSANewsletterFormatter
# ---------------------------------------------------------------------------


class TestAPSANewsletterFormatter:
    def test_format_returns_string(self):
        articles = [_make_apsa_article()]
        md = APSANewsletterFormatter().format(articles, "2026-06")
        assert isinstance(md, str)

    def test_edition_in_header(self):
        md = APSANewsletterFormatter().format([_make_apsa_article()], "2026-06")
        assert "2026-06" in md

    def test_article_title_present(self):
        md = APSANewsletterFormatter().format([_make_apsa_article()], "2026-06")
        assert "AIRWAY EVENTS: THE LATENT GAPS IN PREPAREDNESS" in md

    def test_vignette_in_italics(self):
        md = APSANewsletterFormatter().format([_make_apsa_article()], "2026-06")
        assert "*A high-severity airway event" in md

    def test_body_present(self):
        md = APSANewsletterFormatter().format([_make_apsa_article()], "2026-06")
        assert "Airway events remain one of the most consequential" in md

    def test_references_present(self):
        md = APSANewsletterFormatter().format([_make_apsa_article()], "2026-06")
        assert "Cook TM" in md

    def test_no_references_section_when_empty(self):
        article = _make_apsa_article(references=[])
        md = APSANewsletterFormatter().format([article], "2026-06")
        assert "**References**" not in md

    def test_multiple_articles_all_included(self):
        a1 = _make_apsa_article(title="ARTICLE ONE: SUBTITLE")
        a2 = _make_apsa_article(title="ARTICLE TWO: SUBTITLE")
        md = APSANewsletterFormatter().format([a1, a2], "2026-06")
        assert "ARTICLE ONE" in md
        assert "ARTICLE TWO" in md

    def test_article_count_in_header(self):
        articles = [_make_apsa_article(), _make_apsa_article()]
        md = APSANewsletterFormatter().format(articles, "June 2026")
        assert "Articles:** 2" in md

    def test_footer_verification_note(self):
        md = APSANewsletterFormatter().format([_make_apsa_article()], "2026-06")
        assert "verified before publication" in md

    def test_horizontal_rules_between_articles(self):
        articles = [_make_apsa_article(), _make_apsa_article()]
        md = APSANewsletterFormatter().format(articles, "2026-06")
        assert md.count("---") >= 3
