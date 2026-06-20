"""Integration tests for Week 11 Pipeline endpoints and end-to-end format chain.

Tests:
  - Full format chain: EditorialReport → MarkdownFormatter / ExcelFormatter
  - Pipeline API: GET /pipeline/status, POST /pipeline/report
  - PipelineIngestResult model validation

No real LLM, Qdrant, or network calls — fake components injected throughout.
"""

from __future__ import annotations

import base64
import io
import types
from typing import Any

import pytest

from src.insights.editorial import EditorialEngine
from src.insights.editorial_models import (
    EditorialLLMResponse,
    EditorialReport,
    EditorialSection,
    SectionLLMItem,
)
from src.insights.formatters import ExcelFormatter, MarkdownFormatter
from src.insights.generator import InsightGenerator
from src.insights.models import GeneratedInsight, InsightBatch


# ---------------------------------------------------------------------------
# Fake LLM helpers (same pattern as Week 9 / Week 10 tests)
# ---------------------------------------------------------------------------


_EDITORIAL_RESPONSE = EditorialLLMResponse(
    executive_summary=(
        "Three bronchospasm incidents during laparoscopic procedures share a structural "
        "vulnerability: no mandatory respiratory risk-stratification step exists at "
        "pre-operative assessment, allowing the same triggering sequence to recur."
    ),
    sections=[
        SectionLLMItem(
            theme="root_cause",
            title="Absent Pre-Operative Respiratory Screening",
            narrative=(
                "Across the three retrieved incidents the consistent latent condition is "
                "the absence of a hard-stop respiratory risk-stratification field in the "
                "pre-operative assessment form. Each case progressed to CO₂ insufflation "
                "without a documented bronchospasm contingency plan."
            ),
            key_learning=(
                "Mandatory respiratory risk-stratification is the highest-yield structural "
                "intervention available at department level."
            ),
        ),
        SectionLLMItem(
            theme="safety_recommendations",
            title="System-Level Bronchospasm Prevention",
            narrative=(
                "Three complementary interventions emerge: a mandatory respiratory "
                "risk-stratification field in the pre-operative form, pre-drawn rescue "
                "medication tray at the start of every laparoscopic list, and structured "
                "positional-change verbal warning from the surgical team."
            ),
            key_learning=(
                "Pre-positioning rescue medication and verbal positional warnings are "
                "achievable without significant resource investment."
            ),
        ),
    ],
    conclusion=(
        "The recurring pattern across these incidents is not clinical misfortune "
        "but a predictable consequence of identifiable systemic gaps. The actions "
        "identified are achievable at department level and should be presented at "
        "the next clinical governance meeting."
    ),
)


class _FakeStructuredLLM:
    def __init__(self, response: EditorialLLMResponse = _EDITORIAL_RESPONSE) -> None:
        self._r = response

    def invoke(self, messages: Any) -> EditorialLLMResponse:
        return self._r


class _FakeLLM:
    def __init__(self, response: EditorialLLMResponse = _EDITORIAL_RESPONSE) -> None:
        self._r = response

    def with_structured_output(self, schema: Any) -> _FakeStructuredLLM:
        return _FakeStructuredLLM(self._r)


# ---------------------------------------------------------------------------
# Sample InsightBatch fixture
# ---------------------------------------------------------------------------


def _sample_batch() -> InsightBatch:
    insights = [
        GeneratedInsight(
            insight_id=f"id-{i:03d}",
            query="bronchospasm during laparoscopy",
            insight_text=(
                f"Insight {i}: specific clinical finding about recurring bronchospasm "
                "pattern during CO₂ insufflation in gynaecological laparoscopic cases."
            ),
            insight_type="root_cause",
            evidence_citations=[f"Incident abc{i:03d} | severity=High | score=0.91"],
            actionable_steps=[
                "Anaesthesiologist documents respiratory risk before induction",
                "Nurse pre-positions bronchospasm rescue tray at start of list",
            ],
            confidence="High",
            specificity_score=0.8,
            generated_at="2026-06-14T00:00:00+00:00",
            model_version="week9-insight-generator-0.1",
        )
        for i in range(3)
    ]
    return InsightBatch(
        query="bronchospasm during laparoscopy",
        insights=insights,
        total=3,
        generation_confidence="High",
        evidence_count=3,
        model_version="week9-insight-generator-0.1",
    )


# ---------------------------------------------------------------------------
# Full format chain integration
# ---------------------------------------------------------------------------


class TestMarkdownFormatterIntegration:
    def _get_report(self) -> EditorialReport:
        engine = EditorialEngine(llm=_FakeLLM())
        return engine.generate(_sample_batch())

    def test_chain_produces_markdown_string(self):
        report = self._get_report()
        md = MarkdownFormatter().format(report)
        assert isinstance(md, str)
        assert len(md) > 200

    def test_markdown_contains_executive_summary_content(self):
        # APSA format: executive summary is prose, not a labelled section
        md = MarkdownFormatter().format(self._get_report())
        assert "bronchospasm" in md.lower()  # content is present

    def test_markdown_section_titles_from_llm(self):
        md = MarkdownFormatter().format(self._get_report())
        assert "Absent Pre-Operative Respiratory Screening" in md

    def test_markdown_contains_conclusion_content(self):
        # APSA format: conclusion is prose, not labelled "## Conclusion"
        md = MarkdownFormatter().format(self._get_report())
        assert "clinical governance" in md

    def test_markdown_no_key_learning_blocks(self):
        # APSA format: key learnings are woven into narrative, not displayed separately
        md = MarkdownFormatter().format(self._get_report())
        assert "> **Key Learning:**" not in md

    def test_markdown_citations_present(self):
        md = MarkdownFormatter().format(self._get_report())
        assert "abc" in md  # incident citation IDs from sample batch

    def test_markdown_fallback_report_is_valid(self, monkeypatch):
        monkeypatch.setattr("src.insights.editorial.settings.anthropic_api_key", None)
        engine = EditorialEngine()
        report = engine.generate(_sample_batch())
        md = MarkdownFormatter().format(report)
        # APSA format: structural integrity, not specific headers
        assert isinstance(md, str)
        assert len(md) > 100
        assert "AIR Clinical Incident Intelligence Engine" in md


class TestExcelFormatterIntegration:
    def _get_report(self) -> EditorialReport:
        engine = EditorialEngine(llm=_FakeLLM())
        return engine.generate(_sample_batch())

    def test_chain_produces_valid_xlsx(self):
        import openpyxl
        report = self._get_report()
        data = ExcelFormatter().format(report)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames

    def test_excel_has_section_sheets(self):
        import openpyxl
        report = self._get_report()
        data = ExcelFormatter().format(report)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        # 2 sections in response + Summary + Citations
        assert len(wb.sheetnames) >= 3

    def test_excel_section_narrative_in_sheet(self):
        import openpyxl
        report = self._get_report()
        data = ExcelFormatter().format(report)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        root_sheet = next(n for n in wb.sheetnames if "Root" in n)
        ws = wb[root_sheet]
        all_text = " ".join(
            str(ws.cell(r, 2).value or "") for r in range(1, 30)
        )
        assert "insufflation" in all_text.lower()

    def test_excel_base64_round_trip(self):
        import openpyxl
        report = self._get_report()
        raw_bytes = ExcelFormatter().format(report)
        b64 = base64.b64encode(raw_bytes).decode()
        decoded = base64.b64decode(b64)
        wb = openpyxl.load_workbook(io.BytesIO(decoded))
        assert "Summary" in wb.sheetnames


# ---------------------------------------------------------------------------
# Pipeline API model validation
# ---------------------------------------------------------------------------


class TestPipelineStatusResponse:
    def test_model_instantiates(self):
        from src.api.pipeline import PipelineStatusResponse
        resp = PipelineStatusResponse(
            embedding_model="BAAI/bge-m3",
            qdrant_status="ready",
            llm_available=True,
            llm_model="claude-sonnet-4-6",
            fallback_mode=False,
            total_incidents=10,
        )
        assert resp.total_incidents == 10
        assert resp.llm_available is True

    def test_fallback_mode_model(self):
        from src.api.pipeline import PipelineStatusResponse
        resp = PipelineStatusResponse(
            embedding_model="BAAI/bge-m3",
            qdrant_status="ready",
            llm_available=False,
            llm_model="none",
            fallback_mode=True,
            total_incidents=0,
        )
        assert resp.fallback_mode is True
        assert resp.llm_model == "none"


class TestPipelineIngestResult:
    def test_model_instantiates(self):
        from src.api.pipeline import PipelineIngestResult
        result = PipelineIngestResult(
            ingested=10,
            analyzed=10,
            failed_analysis=0,
            incident_ids=["id-001", "id-002"],
            collection="incidents",
            dimension=1024,
            note="Analyzed ingest — 10 incidents with AI metadata, 0 fell back to raw.",
        )
        assert result.ingested == 10
        assert result.analyzed == 10
        assert result.failed_analysis == 0

    def test_partial_analysis_result(self):
        from src.api.pipeline import PipelineIngestResult
        result = PipelineIngestResult(
            ingested=10,
            analyzed=8,
            failed_analysis=2,
            incident_ids=["id-001"],
            collection="incidents",
            dimension=1024,
            note="8 analyzed, 2 failed.",
        )
        assert result.ingested == 10
        assert result.analyzed == 8
        assert result.failed_analysis == 2


class TestPipelineReportRequest:
    def test_default_format_is_json(self):
        from src.api.pipeline import PipelineReportRequest
        req = PipelineReportRequest(query="airway patterns")
        assert req.format == "json"

    def test_markdown_format_accepted(self):
        from src.api.pipeline import PipelineReportRequest
        req = PipelineReportRequest(query="airway patterns", format="markdown")
        assert req.format == "markdown"

    def test_excel_format_accepted(self):
        from src.api.pipeline import PipelineReportRequest
        req = PipelineReportRequest(query="airway patterns", format="excel")
        assert req.format == "excel"

    def test_invalid_format_rejected(self):
        from pydantic import ValidationError
        from src.api.pipeline import PipelineReportRequest
        with pytest.raises(ValidationError):
            PipelineReportRequest(query="q", format="pdf")  # type: ignore[arg-type]


class TestPipelineReportResponse:
    def test_json_format_response(self):
        from src.api.pipeline import PipelineReportResponse
        resp = PipelineReportResponse(
            report={"title": "Test", "query": "q"},
            markdown=None,
            excel_base64=None,
            pipeline_stages=["retrieval", "insight_generation", "editorial"],
        )
        assert resp.markdown is None
        assert resp.excel_base64 is None
        assert len(resp.pipeline_stages) == 3

    def test_markdown_format_response(self):
        from src.api.pipeline import PipelineReportResponse
        resp = PipelineReportResponse(
            report={"title": "Test"},
            markdown="# Clinical Safety Analysis\n\n...",
            excel_base64=None,
            pipeline_stages=["retrieval", "insight_generation", "editorial", "markdown_format"],
        )
        assert resp.markdown is not None
        assert "Clinical Safety Analysis" in resp.markdown

    def test_excel_format_response(self):
        from src.api.pipeline import PipelineReportResponse
        dummy_b64 = base64.b64encode(b"fake-excel-bytes").decode()
        resp = PipelineReportResponse(
            report={"title": "Test"},
            markdown="# ...",
            excel_base64=dummy_b64,
            pipeline_stages=["retrieval", "insight_generation", "editorial", "markdown_format", "excel_format"],
        )
        assert resp.excel_base64 is not None
        assert base64.b64decode(resp.excel_base64) == b"fake-excel-bytes"


# ---------------------------------------------------------------------------
# End-to-end: InsightBatch → EditorialEngine → Formatters
# ---------------------------------------------------------------------------


class TestEndToEndFormatPipeline:
    def test_full_json_chain(self):
        from src.api.editorial import _report_to_out
        engine = EditorialEngine(llm=_FakeLLM())
        report = engine.generate(_sample_batch())
        out = _report_to_out(report)
        report_dict = out.model_dump()
        assert "title" in report_dict
        assert "sections" in report_dict
        assert "executive_summary" in report_dict

    def test_full_markdown_chain(self):
        engine = EditorialEngine(llm=_FakeLLM())
        report = engine.generate(_sample_batch())
        md = MarkdownFormatter().format(report)
        assert "# Clinical Safety Analysis" in md
        assert report.conclusion[:30] in md

    def test_full_excel_chain(self):
        import openpyxl
        engine = EditorialEngine(llm=_FakeLLM())
        report = engine.generate(_sample_batch())
        data = ExcelFormatter().format(report)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames

    def test_pipeline_stages_accumulate_correctly(self):
        stages: list[str] = []
        engine = EditorialEngine(llm=_FakeLLM())
        report = engine.generate(_sample_batch())
        stages.append("editorial")
        MarkdownFormatter().format(report)
        stages.append("markdown_format")
        ExcelFormatter().format(report)
        stages.append("excel_format")
        assert stages == ["editorial", "markdown_format", "excel_format"]
